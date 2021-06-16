import gc
import openpyxl as pyxl
import pandas as pd
import numpy as np
import win32com.client as win32
from tkinter.ttk import *
from tkinter.filedialog import *
from tkinter import messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
from PIL import ImageTk, Image



class MainApp():

    @staticmethod
    def typeCheck(path):
        pathEle = path.split('.')
        return pathEle.pop().lower()

    @staticmethod
    def readDat(path):
        lists = []

        try:
            with open(path, 'r', encoding="utf-8-sig") as f:
                lines = f.readlines()
            for line in lines:
                line = line.strip('\n')
                line = line.strip('þ')
                lists.append(line.split("þþ"))

            dataLists = lists[1:]
            columnList = lists[0]
            data_array = np.array(dataLists)
            result = pd.DataFrame(dataLists, columns=columnList)
            print("readDat finished")
            return result

        except Exception as e:
            messagebox.showinfo("Warning", "please see the encoding method!(utf-8)")
            print(e)

    @staticmethod
    def changeSlash(path):

        if (path.find("\\") != -1 and path.find("/") == -1):
            path = path.relpace("\\", "/")
            print("changing '/' to '\' completed! ")

        elif (path.find('/') != -1 and path.find("\\") == -1):
            path = path.replace("/", "\\")
            print("changing '\' to '/' completed! ")

        elif (path.find('/') != -1 and path.find('\\') != -1):
            messagebox.showinfo("Warning", "it has both '/' and '\' in its path!")
            print("it has both '/' and '\' in its path!")
            return 0

        return path

    def __init__(self, master):

        self.master = master
        self.master.title("LTS_Create Credential")
        self.master.resizable(width=True, height=True)


        try:
            self.master.iconbitmap(default="./icon/fronteo_98.ico")
            self.imgPath = "./background/bg.png"
            self.img = Image.open(self.imgPath)

            self.img = self.img.resize((1000, 1000), Image.ANTIALIAS)
            self.myImg = ImageTk.PhotoImage(self.img)

            self.label = Label(self.master, image=self.myImg)
            self.label.place(x = 0, y =0, relwidth =1, relheight =1)
        except Exception as e:
            print(e)

        self.mainFrame = Frame(master, relief="raised")
        self.mainFrame.pack(padx=70, pady=10, expand=True, anchor = "n")


        self.loadObj = LoadOption(self)
        self._pathInfo = dict()             # 파일 위치 정보 - 딕셔너리
        self._accountData = pd.DataFrame()  # 계정 정보 DataFrame
        self._writeOptDict = dict()  # 사용자 입력 사항 - 딕셔너리
        self._saveOpt = []  # Save option - 리스트
        self.writeObj = None
        self.saveObj = None
        print("init Object Count: ", end = "")
        print(len(gc.get_objects()))



    # getter & setter 함수 지정

    # 1. _pathInfo: getter & setter
    @property
    def pathInfo(self):
        return self._pathInfo

    @pathInfo.setter
    def pathInfo(self, value):  # value는 dictionary
        self._pathInfo.clear()
        self._pathInfo.update(value)

    # 2. _accountInfo : only getter

    @property
    def accountData(self):
        return self._accountData

    # 3. _writeOptDict

    @property
    def writeOptDict(self):
        return self._writeOptDict

    @writeOptDict.setter
    def writeOptDict(self, value):
        self._writeOptDict.clear()
        self._writeOptDict.update(value)

    # 4. _saveOpt

    @property
    def saveOpt(self):
        return self._saveOpt

    @saveOpt.setter
    def saveOpt(self, value):
        self._saveOpt.clear()
        self._saveOpt.extend(value)

    # LoadOption Class(self.loadObj)의 Read button과 binding
    # LoadOption Class의 readPathInfo에서 Call

    def loadData(self):

        pathInfoKeys = list(self._pathInfo.keys())
        onlyPathKeys = pathInfoKeys[0:2]

        try:
            self._accountData = pd.DataFrame()  # 계정 정보 DataFrame
            # \, / 변환 기능
            for key in onlyPathKeys:
                self._pathInfo[key] = MainApp.changeSlash(self._pathInfo[key])

            # 파일 타입별 처리 분기.

            ## 리팩토링 필요 => typeCheck 함수와 fileType check를 한번에?

            if (self._pathInfo["fileType"] == 0):  # Concordance / utf-16
                if (MainApp.typeCheck(self._pathInfo["accountPath"]) != "dat"):
                    messagebox.showinfo("Warning", "file type is not dat!")
                    return 1
                self._accountData = MainApp.readDat(self._pathInfo["accountPath"])
                print("accountData has been loaded: ")
                print(self._accountData)
            elif (self._pathInfo["fileType"] == 1):  # CSV
                if (MainApp.typeCheck(self._pathInfo["accountPath"]) != "csv"):
                    messagebox.showinfo("Warning", "file type is not csv!")
                    return 1
                self._accountData = pd.read_csv(self._pathInfo["accountPath"], encoding="utf-8")
                print("accountData has been loaded: ")
                print(self._accountData)
            elif (self._pathInfo["fileType"] == 2):  # Excel
                if (MainApp.typeCheck(self._pathInfo["accountPath"]) != "xlsx"):
                    messagebox.showinfo("Warning", "file type is not xlsx!")
                    return 1
                self._accountData = pd.read_excel(self._pathInfo["accountPath"])
                print("accountData has been loaded: ")
                print(self._accountData)
            else:
                print("File type error!")
                messagebox.showinfo("Warning", "File type error!")

            if (MainApp.typeCheck(self._pathInfo["excelPath"]) != "xlsx"):
                messagebox.showinfo("Warning", "excel form file type can only use 'xlsx'")
                return 1

            self._wb = pyxl.load_workbook(self._pathInfo["excelPath"], data_only=True)
            print("Excel Workbook has been loaded successfully!")
            messagebox.showinfo("Info", "All the data has been loaded successfully!")
            return 0

        except Exception as e:
            print(e)


    # Load Option class의 Clear button과 binding

    def clear (self, event):
        try:
            self._wb.close()
            self._pathInfo.clear()
            del self._accountData
            self._writeOptDict.clear()
            self._saveOpt.clear()

            if (not self.writeObj):
                print("self.writeOjb is empty")
                pass
            else:
                self.writeObj.__del__()

            if (not self.saveObj):
                print("self.saveObj is empty")
                pass
            else:
                self.saveObj.__del__()

        except Exception as e:
            print(e)

        finally:
            self._pathInfo = dict()  # 파일 위치 정보 - 딕셔너리
            self._accountData = pd.DataFrame()  # 계정 정보 DataFrame
            self._writeOptDict = dict()  # 사용자 입력 사항 - 딕셔너리
            self._saveOpt = []  # Save option - 리스트
            self.writeObj = None
            self.saveObj = None
            print("모든 메모리 초기화")
            print("check the object count before garbage collect: ", end = "")
            print(len(gc.get_objects()))
            gc.collect()
            print("check the object count after garbage collect: ", end = "")
            print(len(gc.get_objects()))



    # Call Write Option Class(self.WriteOption) to MainApp
    # LoadOption Class(self.loadObj)의 Read button과 binding
    # LoadOption Class의 readPathInfo에서 Call


# ========================================================== self._wbSheets 와 self._accColList 변수 관리.
    def callWriteOption(self):
        self._wbSheets = list(self._wb.sheetnames)
        print("workbook's sheet list: ")
        print(self._wbSheets)
        self._accColList = list(self._accountData.columns)
        print("account Info file's column : ")
        print(self._accColList)
        self.writeObj = WriteOption(self, self._accColList, self._wbSheets)
        print("WriteOption object(Widget) has been loaded!")

    def callSaveOption(self):
        accColList = list(self._accountData.columns)
        print("account Info file's column : ", end="")
        print(accColList)
        self.saveObj = SaveOption(self, accColList)

    def dropPath(self, event):
        event.widget.delete(0, "end")
        path = event.data
        # print(event.widget);

        try:
            if (path[0] == "{"):
                path = path.strip('{}')
            print("입력 path = ", end="")
            print(path)

        except Exception as e:
            print(e)
        event.widget.insert("end", path)

    def saveCredential(self):

        try:
            for i in range(len(self._accountData)):
                print("i: " + str(i))
                for accCol in self._writeOptDict.keys():
                    print("accCol: " + str(accCol))
                    for sheets in self._writeOptDict[accCol].keys():
                        print("sheets: " + str(sheets))
                        cellList = (self._writeOptDict[accCol][sheets].split(","))
                        ws = self._wb[self._wbSheets[sheets]]

                        for cell in cellList:
                            cell = cell.strip(" ")
                            print(self._accountData)
                            print(self._accountData[self._accColList[accCol]][i])
                            ws[cell] = self._accountData[self._accColList[accCol]][i]
                save_name_s = self.saveOpt[0].format(self._accountData[self.saveOpt[1][0]][i])
                self._wb.save(save_name_s)
                self._wb.close()

                if (self.saveOpt[1][1] == "Disabled" or self.saveOpt[1][1] == "UnSelected"):
                    continue
                # password setting.
                save_name_bs = MainApp.changeSlash(save_name_s)
                print("save path back slash: ", end="")
                print(save_name_bs)
                try:
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    excel.Visible = True
                    excel.DisplayAlerts = False
                    print(excel)
                    if (i == 0):
                        wb2 = excel.Workbooks.Open(save_name_bs)
                        wb2.SaveAs(save_name_bs, 51, self._accountData[self.saveOpt[1][1]][i])
                    else:
                        wb2 = excel.Workbooks.Open(save_name_bs, Password=self._accountData[self.saveOpt[1][1]][0])
                        wb2.SaveAs(save_name_bs, 51, self._accountData[self.saveOpt[1][1]][i])

                    print(str(i) + "번째의 " + "password decrption Working Good")
                    wb2.Close()
                    excel.Application.Quit()


                except Exception as e:
                    print(e)
                    wb2 = None
                    excel.Application.Quit()
                    return 1

        except Exception as e:
            print(e)
            return 1

        return 0

    # ========================================= class 정의  ===========================================


class LoadOption():

    def __init__(self, mainApp):

        self.mainApp = mainApp
        self.loadFrame = LabelFrame(self.mainApp.mainFrame, text="Load Option", relief = "sunken")
        self.loadFrame.pack(padx=10, pady=3, fill="x", side="top")

        # pack으로 했는데도 grid_columnconfigure 적용 가능함.

        self.loadFrame.grid_columnconfigure(0, weight=3)
        self.loadFrame.grid_columnconfigure(1, weight=1)

        """
        path Entry and Browse Frame ===================================================== 
        """
        self.pathFrame = Frame(self.loadFrame)
        self.pathFrame.grid(row=0, column=0, sticky="NSWE", padx=3)

        self.pathFrame.grid_columnconfigure(0, weight=1, uniform="---")
        self.pathFrame.grid_columnconfigure(1, weight=2, uniform="---")

        # Label for Account Entry
        self.accountLabel = Label(self.pathFrame, text="Account Info File Full path")
        self.accountLabel.grid(row=0, column=0, columnspan=2, padx=5, ipady=5, sticky="w")

        # Account Info Entry
        self.accountInfoEntry = Entry(self.pathFrame)
        self.accountInfoEntry.grid(padx=5, pady=5, ipady=3, row=1, column=0, sticky="WE", columnspan=2)

        # Drag & Drop function
        self.accountInfoEntry.drop_target_register(DND_FILES)
        self.accountInfoEntry.dnd_bind("<<Drop>>", self.mainApp.dropPath)

        # Browse Button - Callback function binding
        self.accountBrowse = Button(self.pathFrame, relief="raised", text="Browse", name="accBrowse")
        self.accountBrowse.grid(row=1, column=7, padx=5)
        self.accountBrowse.bind("<ButtonRelease-1>", self.browseCallback)

        # Label for Excel Entry
        self.excelLabel = Label(self.pathFrame, text="Excel Form File Full path")
        self.excelLabel.grid(row=2, column=0, padx=5, ipady=5, sticky="w")

        # Excel Form Entry
        self.excelFormEntry = Entry(self.pathFrame)
        self.excelFormEntry.grid(row=3, column=0, padx=5, pady=5, ipady=3, sticky="WE", columnspan=2)

        # Drag & Drop function
        self.excelFormEntry.drop_target_register(DND_FILES)
        self.excelFormEntry.dnd_bind("<<Drop>>", self.mainApp.dropPath)

        # Browse Button and Callback function binding
        self.excelBrowse = Button(self.pathFrame, relief="raised", text="Browse", name="excelBrowse")
        self.excelBrowse.grid(row=3, column=7, padx=5)
        self.excelBrowse.bind("<ButtonRelease-1>", self.browseCallback)

        """
        Path Frame End ================================================================ 

        File Type Frame Start
        """
        # print(str(self.excelBrowse)) // button widget naming test 완료.

        # File Type Option Frame
        self.optionFrame = LabelFrame(self.loadFrame, text="File Type")
        self.optionFrame.grid(row=0, column=1, sticky="NSWE", padx=3)

        # File Type RadioButton

        self.fileTypeVar = IntVar()

        self.datOption = Radiobutton(self.optionFrame, text="Concordance", value=0, variable=self.fileTypeVar)
        self.datOption.pack(anchor="w")
        self.csvOption = Radiobutton(self.optionFrame, text="CSV(UTF-8)", value=1, variable=self.fileTypeVar)
        self.csvOption.pack(anchor="w")
        self.excelOption = Radiobutton(self.optionFrame, text="Excel", value=2, variable=self.fileTypeVar)
        self.excelOption.pack(anchor="w")

        # Read & Clear Option Frame
        self.readFrame = Frame(self.loadFrame, relief = "sunken")
        self.readFrame.grid(row=0, column=3, sticky="NSWE", padx=3)

        """
        self.testButton2 = Button(self.readFrame, text = "test", width = 10)
        self.testButton2.pack()
        """
        self.readButton = Button(self.readFrame, text="read", width=10)
        self.readButton.pack(side="top", pady=15)
        self.readButton.bind("<ButtonRelease-1>", self.readPathInfo)

        self.clearButton = Button(self.readFrame, text="clear", width=10)
        self.clearButton.pack(side="bottom", pady=15)
        self.clearButton.bind("<ButtonRelease-1>", self.mainApp.clear)

    # tkinter event.key => tkinterDnD event.widget (event가 발생한 widget 호출)

    def dropPath(self, event):
        event.widget.delete(0, "end")
        path = event.data
        # print(event.widget);

        try:
            if (path[0] == "{"):
                path = path.strip('{}')
            print("입력 path = ", end="")
            print(path)

        except Exception as e:
            print(e)
        event.widget.insert("end", path)

    ## Brwose Button Callback 함수

    def browseCallback(self, event):
        # get widget name

        eWidget = str(event.widget)
        eWidget = eWidget.split(".")
        print("browseCallback이 눌린 widget: ", end="")
        print(eWidget[-1])

        path = askopenfilename(initialdir="C:/Users", title="Find Account Info", filetypes=(
        ("excel files", "*.xlsx"), ("csv files", "*.csv"), ("dat files", "*.dat"), ("all files", "*.*")))

        if path == "":
            print("path unchecekd")
            return 0

        if (eWidget[-1] == "accBrowse"):
            try:
                self.accountInfoEntry.delete(0, "end")
                self.accountInfoEntry.insert("end", path)
                print("Account Info Path: ", self.accountInfoEntry.get())

            except Exception as e:
                messagebox.showinfo("Warning", e)

        elif (eWidget[-1] == "excelBrowse"):
            try:
                self.excelFormEntry.delete(0, "end")
                self.excelFormEntry.insert("end", path)
                print("Excel Form Path: ", self.excelFormEntry.get())

            except Exception as e:
                messagebox.showinfo("Warning", e)
        else:
            messagebox.showinfo("Warning", "Unkown Error")

    def readPathInfo(self, event):

        try:
            if (not (self.mainApp.accountData.empty) or list(self.mainApp.accountData.columns)):
                print(self.mainApp.accountData)
                messagebox.showinfo("Warning", "Data has been uploaded, Please click the clear before reload it")
                return 1

            tempDict = {"accountPath": self.accountInfoEntry.get(), "excelPath": self.excelFormEntry.get(),
                        "fileType": self.fileTypeVar.get()}
            print("tempDict의 items for QC : ", end="")
            print(tempDict.items())
            self.mainApp.pathInfo = tempDict
            print("mainApp의 pathInfo for QC : ", end="")
            print(self.mainApp.pathInfo.items())

            self.mainApp.loadData()
            self.mainApp.callWriteOption()

        except Exception as e:
            messagebox.showinfo("Warning", e)



# ================================================ WriteOption Class ==========================

class WriteOption():


    def __init__(self, mainApp, accColList, excelSheets):
        self.mainApp = mainApp
        self.accColList = accColList
        self.excelSheets = excelSheets

        # self.loadFrame = LabelFrame(self.mainApp.mainFrame, relief = "solid", bd = 1, bg = "yellowgreen", text = "Load Option")
        self.writeFrame = LabelFrame(self.mainApp.mainFrame, relief="sunken", text="Write Option")
        self.writeFrame.pack(padx=10, pady=3, fill="x", side="top")

        self.canvasFrame = Frame(self.writeFrame)
        self.canvasFrame.pack(side="top", fill="x", padx=10)

        # canvasObj
        self.canvasObj = ScrollCanvas(self.canvasFrame, self.accColList, self.excelSheets)

        self.applyButton = Button(self.writeFrame, text="Apply", width=10)
        self.applyButton.pack(anchor="e", padx=15, pady=10, side="top")
        self.applyButton.bind("<ButtonRelease-1>", self.applyWriteOption)


    def __del__(self):
        try:
            for child in self.writeFrame.winfo_children():
                child.destroy()
            self.writeFrame.destroy()
        except Exception as e:
            print(e)
        finally:
            gc.collect()
            print("WriteOption widget has been deleted!")


    # Apply Button을 눌렀을 때 호출될 함수

    def applyWriteOption(self, event):

        if(not self.mainApp.saveObj):
            print("SaveObj is empty!")
        else:
            self.mainApp.saveObj.__del__()
            self.mainApp.saveObj = None
            print(len(gc.get_objects()))
            gc.collect()
            print(len(gc.get_objects()))

        transferDict = dict()

        for insObj in self.canvasObj._insertObjList:
            sheetDict = dict()
            if (insObj.myCheckVar.get() == 0):
                continue;
            else:
                for sheetIdx, sheetEntry in enumerate(insObj._entryList):
                    if (not sheetEntry.get()):
                        continue;
                    sheetDict[sheetIdx] = sheetEntry.get()
                transferDict[insObj.idx] = sheetDict
                print(insObj.idx, end=" : ")
                print(transferDict[insObj.idx])

        self.mainApp.writeOptDict = transferDict;
        print(self.mainApp.writeOptDict)

        self.mainApp.callSaveOption()

        return 0


# ---------------------- Scrollerable Canvas class --------------------------

class ScrollCanvas(Frame):

    def __init__(self, canvasFrame, accColList, excelSheets):

        # arguments initialization

        self.canvasFrame = canvasFrame
        self.accColList = accColList
        self.excelSheets = excelSheets

        # insert Objecet를 담을 List
        self._insertObjList = []

        self.canvas = Canvas(self.canvasFrame)
        self.frame = Frame(self.canvas)
        self.vsb = Scrollbar(self.canvasFrame, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="top", fill="x", expand=True)
        self.canvas_frame = self.canvas.create_window((4, 4), window=self.frame, anchor="center", tags="self.frame")

        # 작성될 frame에 widget 추가 될 때마다, yscroll이 반응할 수 있게 바인딩.
        self.frame.bind("<Configure>", self.onFrameConfigure)
        self.canvas.bind("<Configure>", self.FrameWidth)

        self.populate()

        if (len(self._insertObjList) == len(self.accColList)):
            print("모든 열 객체화 생성 성공!")
        else:
            messagebox.showinfo("Warning", "Column loading fail")

    #  populate the InsertObj

    def populate(self):
        try:
            for i in range(len(self.accColList)):
                self._insertObjList.append(InsertObj(i, self))
        except Exception as e:
            messagebox.showinfo("Warning", e)

    def onFrameConfigure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    # Canvas의 Frame width를 Canvas만큼 최대한 확장 기능.
    def FrameWidth(self, event):
        canvas_width = event.width
        # print(canvas_width)
        self.canvas.itemconfig(self.canvas_frame, width=canvas_width)


# ------------------------  Insert Object -----------------------------------------

class InsertObj():

    def __init__(self, idx, canvasObj):

        # arguments initialization.
        self.idx = idx
        self.canvasObj = canvasObj

        # Insert Object Frame 생성

        self.InsertFrame = Frame(self.canvasObj.frame, relief = "sunken", bd = 1)
        self.InsertFrame.pack(fill="x", padx = 3, pady = 3)

        self.InsertFrame.grid_columnconfigure(0, weight=3, uniform="---")
        self.InsertFrame.grid_columnconfigure(1, weight=7, uniform="---")

        # Checkbox 설정
        self.accColFrame = Frame(self.InsertFrame)
        self.accColFrame.grid(row=0, column=0, sticky="nswe", padx=10)

        self.myCheckVar = IntVar()
        self.myCheckButton = Checkbutton(self.accColFrame, text=self.canvasObj.accColList[self.idx],
                                         variable=self.myCheckVar, command=self.enableEntry)
        self.myCheckButton.pack(padx=10, side="bottom", pady=5, anchor="w")

        # Excel Position Frame

        self.excelPosFrame = Frame(self.InsertFrame)
        self.excelPosFrame.grid(row=0, column=1, sticky="nswe")

        self.excelNotebook = Notebook(self.excelPosFrame)
        self.excelNotebook.pack(padx=20, pady=5, anchor="w", fill="x")

        # excel sheet별 frame과 entry 저장.
        self._frameList = []
        self._entryList = []

        for i in range(len(self.canvasObj.excelSheets)):
            self._frameList.append(Frame(self.excelPosFrame))
            self._entryList.append(Entry(self._frameList[i], state="disabled"))
            self._entryList[i].pack(fill="x")
            self.excelNotebook.add(self._frameList[i], text=self.canvasObj.excelSheets[i])


    def enableEntry(self):
        if (self.myCheckVar.get() == 0):
            print(self.myCheckVar.get())
            for entry in self._entryList:
                entry["state"] = "disabled";
        else:
            print(self.myCheckVar.get())
            for entry in self._entryList:
                entry["state"] = "normal";


class SaveOption():


    def __init__(self, mainApp, accColList):
        self.mainApp = mainApp
        self.accColList = accColList
        self.saveFrame = LabelFrame(self.mainApp.mainFrame, text="Save Option", relief = "sunken")
        self.saveFrame.pack(padx=10, pady=3, fill="x", side="top")

        self.formFrame = Frame(self.saveFrame, relief = "ridge", bd = 1)
        self.formFrame.pack(fill="both", padx=3, pady=3)

        self.formFrame.grid_columnconfigure(0, weight=1)
        self.formFrame.grid_columnconfigure(1, weight=1)
        self.formFrame.grid_columnconfigure(2, weight=1)

        self.formLabel = Label(self.formFrame, text="Insert Form")
        self.formLabel.grid(row=0, column=0, sticky="ns", padx=10)
        self.formEntry = Entry(self.formFrame)
        self.formEntry.grid(row=0, column=1, sticky="nsew", columnspan=2)

        self.pathLabel = Label(self.formFrame, text="Save Path")
        self.pathLabel.grid(row=1, column=0, sticky="ns", padx=10)
        self.pathEntry = Entry(self.formFrame)
        self.pathEntry.grid(row=1, column=1, sticky="nswe", columnspan=2)

        self.pathEntry.drop_target_register(DND_FILES)
        self.pathEntry.dnd_bind("<<Drop>>", self.mainApp.dropPath)

        # Bottom Frame 시작.

        self.bottomFrame = Frame(self.saveFrame, relief = "ridge")
        self.bottomFrame.pack(fill="both")

        self.bottomFrame.grid_columnconfigure(0, weight=1)
        self.bottomFrame.grid_columnconfigure(1, weight=1)
        self.bottomFrame.grid_columnconfigure(2, weight=1)

        self.bottomFrame.grid_rowconfigure(0, weight=1)
        self.bottomFrame.grid_rowconfigure(1, weight=1)
        self.bottomFrame.grid_rowconfigure(2, weight=1)

        ## Naming Frame

        self.namingFrame = Frame(self.bottomFrame, relief = "raised", bd = 1)
        self.namingFrame.grid(row=0, column=0, sticky="nswe", padx=5)

        self.selectLabel = Label(self.namingFrame, text="select field to insert")
        self.selectLabel.pack(side="top", fill="both")

        self.listFrame = Frame(self.namingFrame)
        self.listFrame.pack(fill="both")

        self.n_scrollbar = Scrollbar(self.listFrame)
        self.n_scrollbar.pack(side="right", fill="y")

        self.n_fieldListBox = Listbox(self.listFrame, selectmode="single", yscrollcommand=self.n_scrollbar.set)
        self.n_fieldListBox.pack(fill="both")
        self.n_fieldListBox.bind("<ButtonRelease-1>", self.printSelected)

        self.n_scrollbar.config(command=self.n_fieldListBox.yview)

        for col in accColList:
            self.n_fieldListBox.insert(END, col)

        ## Password Frame 시작

        self.pwFrame = Frame(self.bottomFrame, relief = "raised", bd =1)
        self.pwFrame.grid(row=0, column=1, sticky="nswe", padx=5)

        self.pwCheckVar = IntVar()
        # self.myCheckButton = Checkbutton(self.accColFrame, text = self.canvasObj.accColList[self.idx], variable = self.myCheckVar, command = self.enableEntry)
        self.pwCheckButton = Checkbutton(self.pwFrame, text="select field for PW", variable=self.pwCheckVar,
                                         command=self.enablePW)
        self.pwCheckButton.pack(side="top", fill="x")

        self.pwListFrame = Frame(self.pwFrame)
        self.pwListFrame.pack(fill="both")

        self.p_scrollbar = Scrollbar(self.pwListFrame)
        self.p_scrollbar.pack(side="right", fill="y")

        self.p_fieldListBox = Listbox(self.pwListFrame, selectmode="single", yscrollcommand=self.p_scrollbar.set)
        self.p_fieldListBox.pack(fill="both")

        self.p_scrollbar.config(command=self.p_fieldListBox.yview)

        for col in accColList:
            self.p_fieldListBox.insert(END, col)

        self.p_fieldListBox['state'] = "disabled"
        self.p_fieldListBox.bind("<ButtonRelease-1>", self.printSelected)

        # statusFrame 시작

        self.statusFrame = Frame(self.bottomFrame, relief = "raised", bd =1)
        self.statusFrame.grid(row=0, column=2, sticky="nswe")

        self.statusFrame.grid_columnconfigure(0, weight=1)
        self.statusFrame.grid_columnconfigure(1, weight=1)
        self.statusFrame.grid_columnconfigure(2, weight=1)

        self.statusFrame.grid_rowconfigure(0, weight=1)
        self.statusFrame.grid_rowconfigure(1, weight=5)
        self.statusFrame.grid_rowconfigure(2, weight=1)

        self.statusLabel = Label(self.statusFrame, text="Status Info")
        self.statusLabel.grid(row=0, column=0, sticky="NSWE", columnspan=3)

        self.statusList = ["Unselected", "Disabled"]

        self.textVariable = "Naming Format: {0}\nPassword Format: {1}".format(self.statusList[0], self.statusList[1])

        self.infoLabel = Label(self.statusFrame, text=self.textVariable)
        self.infoLabel.grid(row=1, column=0, pady=10, sticky="NSWE", columnspan=3, rowspan=2)
        self.infoLabel.bind("<ButtonRelease-1>", self.printSelected)

        self.saveButton = Button(self.statusFrame, text="save", width=10)
        self.saveButton.grid(row=3, column=0, padx=10, pady=10, columnspan=3)
        self.saveButton.bind("<ButtonRelease-1>", self.save)

    def __del__(self):
        try:
            for child in self.saveFrame.winfo_children():
                child.destroy()
            self.saveFrame.destroy()

        except Exception as e:
            print(e)
        finally:
            gc.collect()
            print("saveFrame widget has been deleted!")

    def enablePW(self):
        if (self.pwCheckVar.get() == 0):
            self.p_fieldListBox['state'] = "disabled"
            self.statusList[1] = "Disabled"
        else:
            self.p_fieldListBox['state'] = "normal"
            self.statusList[1] = "UnSelected"
        self.textVariable = "Naming Format: {0}\nPassword Format: {1}".format(self.statusList[0], self.statusList[1])
        self.infoLabel["text"] = self.textVariable

    def printSelected(self, event):

        eWidget = str(event.widget)
        print(eWidget)
        eWidget = eWidget.split(".")
        print("선택된 column ListBox widget: ", end="")
        print(eWidget[-3])

        if (eWidget[-3] == "!frame"):
            self.statusList[0] = event.widget.get(event.widget.curselection())
        else:
            self.statusList[1] = event.widget.get(event.widget.curselection())

        print(event.widget.get(event.widget.curselection()))
        self.textVariable = "Naming Format: {0}\nPassword Format: {1}".format(self.statusList[0], self.statusList[1])
        self.infoLabel["text"] = self.textVariable

    def save(self, event):

        if (not not self.pathEntry.get()):

            if (self.pathEntry.get()[-1] != "/"):
                self.pathEntry.insert(END, "/")

        pathForm = self.pathEntry.get() + self.formEntry.get()
        print(pathForm)

        saveOptTemp = [pathForm, self.statusList]
        print(saveOptTemp)
        self.mainApp.saveOpt = saveOptTemp
        print(self.mainApp.saveOpt)

        # write excel function call

        self.mainApp.saveCredential();


# ================================= main 함수 ====================================

def main():

    root = TkinterDnD.Tk()
    scrWidth = root.winfo_screenwidth()
    scrHeight = root.winfo_screenheight()
    print("%sx%s" % (int(scrWidth / 2), int(scrHeight * 4 / 5)))

    root.geometry("%sx%s" % (int(scrWidth / 2), int(scrHeight * 4 / 5)))
    root.minsize(960, 800)

    app = MainApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()

